#!/usr/bin/python2

from __future__ import division
from __future__ import print_function
import argparse
import os
import csv
from time import sleep
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy,deepcopy

def transform_and_sheet_split(file_name):
    '''Transform all sheet and seperate the file into csv files
    sheet by sheet'''

    # Read file
    try:
        work_book = load_workbook(filename=file_name)
        pass

    except IOError as e:
        print("Error: No such file, please check out the file name")

    else:
        # Read by columns and write by rows 
        sheet_names = work_book.get_sheet_names()
        for temp_sheet_name in sheet_names:
            with open(file_name[:-5]+'_'+temp_sheet_name+'.csv','wb') as tempfile:
                tempwriter = csv.writer(tempfile)
                temp_sheet = work_book[temp_sheet_name]
                for col in temp_sheet.iter_cols(values_only=True):
                    tempwriter.writerow(col)

        print('File '+file_name+' Processed Successfully')


def average_per_six_columns(file_name):
    '''Get the origin data in the xlsx file and calculate the 
    average of target data in the rule of every six data'''

    # Read file
    try:
        work_book = load_workbook(filename=file_name)
        pass

    except IOError as e:
        print("Error: No such file, please check out the file name")

    else:
        # Read by columns and write by rows 
        sheet_names = work_book.get_sheet_names()
        for temp_sheet_name in sheet_names:
            
            temp_sheet = work_book[temp_sheet_name]

            for x in xrange((temp_sheet.max_column-1)//6):

                temp_sheet.insert_cols(7*(x+1)+1)

                temp_cell = temp_sheet.cell(row=1, column=7*(x+1)+1)
                temp_cell.value = 'Average'
                left_cell = temp_sheet.cell(row=1, column=7*(x+1))
                if left_cell.has_style:
                    temp_cell._style = copy(left_cell._style)

                for row, cellObj in enumerate(list(temp_sheet.columns)[7*(x+1)]):
                    if row:
                        cellObj.value  = '=AVERAGE(%s%d:%s%d)' % (get_column_letter(7*x+2), row+1, get_column_letter(7*(x+1)), row+1)

        output_name = file_name
        if output_name[0] == '.':
            output_name = file_name.split('/', 1)[1]

        output_name = './output/'+output_name
        if not os.path.exists(os.path.dirname(output_name)):
            try:
                os.makedirs(os.path.dirname(output_name))
                pass
            except Exception as e:
                raise e

        work_book.save(output_name)
        print('File '+file_name+' Processed Successfully')


function_list = [transform_and_sheet_split, average_per_six_columns]


def folder_process(folder_address, function_type):
    '''Find out all the xlsx files in the folder with 
    going into all the sub-directories, and process'''

    file_list = []
    list_dirs = os.walk(folder_address)
    for root, dirs, files in list_dirs:
        for f in files:
            try:
                if f[-5:] == ".xlsx" or f[-5:] == ".XLSX":
                    file_list.append(os.path.join(root, f))
            except Exception as e:
                raise e

    pbar = tqdm(total=len(file_list),dynamic_ncols=True,unit='file')

    for temp_file in file_list:
        pbar.set_description(desc='Processing: '+temp_file)
        file_process(temp_file, function_type)
        sleep(0.01)
        pbar.update(1)
        
    pbar.close()


def file_process(file_name, function_type):
    '''Process with particular file'''

    function_list[function_type-1](file_name)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='excelprocess')
    parser.add_argument('--file', type=str, default='test.xlsx', help='determine which excel file to be processed, for example: extract.py --file example.xlsx')
    parser.add_argument('--folder', type=str, help='determine the folder where all the xlsx files will to be processed, for example: extract.py --folder ./address')
    parser.add_argument('--function', type=int, default=1, help='determine what to do on the original files, \nfunctions include: 1 - transform, sheet_split 2 - average per six columns')

    arg = parser.parse_args()

    if arg.function<1 or arg.function>2 :
        print("Function parameter error!\nPlease refer to help")
        raise SystemExit

    if arg.folder:
        folder_process(arg.folder, arg.function)
    else:
        file_process(arg.file, arg.function)